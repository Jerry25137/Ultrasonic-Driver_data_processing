# -*- coding: utf-8 -*-
"""
Update：2025.05.23

Version：3.1.4 Engineering Edition

@author: Hsiao Yu-Chieh
"""

import os
import csv
import time

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties
import win32com.client as win32

import matplotlib.pyplot as plt
import matplotlib
matplotlib.rcParams['font.family'] = 'Microsoft JhengHei' # 設定中文字體
matplotlib.use('Qt5Agg')                                  # 指定互動式後端

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

version = "EE"

# 主程式模組--------------------------------------------------------------------------------------------------

# 11種線型
linetypes = ["solid", "sysDash", "sysDashDot", "sysDashDotDot", "sysDot",
             "dash", "dashDot", "dot", "lgDash", "lgDashDot", "lgDashDotDot",]

# 54種循環顏色
colors = ["4F81BD", "C0504D", "9BBB59", "8064A2", "4BACC6", "F79646", "2C4D75", "772C2A", "5F7530", "4D3B62", 
          "276A7C", "B65708", "729ACA", "CD7371", "AFC97A", "9983B5", "6FBDD1", "F9AB6B", "3A679C", "9F3B38", 
          "7E9D40", "664F83", "358EA6", "F3740B", "95B3D7", "D99694", "C3D69B", "B3A2C7", "93CDDD", "FAC090", 
          "254061", "632523", "4F6228", "403152", "215968", "984807", "84A7D1", "D38482", "B9CF8B", "A692BE", 
          "81C5D7", "F9B67E", "335A88", "8B3431", "6F8938", "594573", "2E7C91", "D56509", "A7C0DE", "DFA8A6", 
          "CDDDAC", "BFB2D0", "A5D6E2", "FBCBA3", ]

# 讀取檔案 (路徑, 檔案類型)
def get_files_in_dir(f_path):
    # 取得目前目錄中的所有檔案名
    files = os.listdir(f_path)
    
    # 篩選出.csv或.txt 結尾的檔名，並將它們儲存到清單中
    files_csv = [filename for filename in files if filename.endswith(".csv") or filename.endswith(".CSV")]
    files_txt = [filename for filename in files if filename.endswith(".txt") or filename.endswith(".TXT")]

    # 合併.csv和.txt清單
    files = files_csv + files_txt

    return files

# UD系列資料處理(路徑)
def UD_data(f_path):
    UD2_Current = []
    UD2_Phase   = []
    UD7_Current = []
    
    # 取得文件名列表
    file_list = get_files_in_dir(f_path)

    # 用來追蹤是否有補充資料
    data_filled = False
    
    try:
        for f in file_list:
            if f[-3:] == "csv" or f[-3:] == "CSV":
                with open(os.path.join(f_path, f), 'r', newline = '', encoding = 'utf-8-sig') as file:
                    reader = csv.reader(file)
                    
                    data_title = next(reader)
                    if data_title == ['Freq', 'Current', 'Phase', ''] or data_title == ['freq', 'Current']:
                        rows = [ [int(item) if item.isdigit() else item for item in row if item.strip()]
                                 for row in reader
                                 if row  ] # 過濾掉空行
                    
                    else:
                        continue
                    
            elif f[-3:] == "txt" or f[-3:] == "TXT":
                file_path = os.path.join(f_path, f)

                # 讀取txt檔案
                with open(file_path, 'r', encoding = 'utf-8-sig') as file:
                    reader = file.readlines()

                if "," in reader[0] and reader[0] == "Freq,Current,Phase,\n":
                    # 按逗號分割，去除空行，清理並將數字轉換為 int
                    rows = [ [int(item.strip()) if item.strip().isdigit() else item.strip() 
                             for item in line.split(",") if item.strip()]
                             for line in reader
                             if line.strip() ]
                
                elif "	" in reader[0] and reader[0] == "Freq - Plot 0\tCurrent - Plot 0\n":
                    # 按照空白字符分割每行內容並存入列表
                    orig_rows = [ [int(item) if item.isdigit() else item
                             for item in line.strip().split('\t')]
                             for line in reader
                             if line.strip() ] # 過濾空行

                    header = orig_rows[0]          # 擷取資料開頭
                    body   = orig_rows[1:]         # 擷取資料數據
                    rows   = [header] + body[::-1] # 合併開頭與數據
                
                else:
                    continue

            # 將資料替換成檔名作為標籤
            rows[0][0] = "Freq"
            rows[0][1] = os.path.splitext(f)[0] # 電流替換成檔名作為標籤
            
            # UD2
            if len(rows[0]) == 3:
                rows[0][2] = os.path.splitext(f)[0] # 相位替換成檔名作為標籤
                UD2_Current_temp = [] # 電流暫存資料
                UD2_Phase_temp   = [] # 相位暫存資料
                
                # 建立暫存資料
                for i1 in range(len(rows)):
                    UD2_Current_temp.append([])
                    UD2_Current_temp[i1].append(rows[i1][0])
                    UD2_Current_temp[i1].append(rows[i1][1])
                    
                    UD2_Phase_temp.append([])
                    UD2_Phase_temp[i1].append(rows[i1][0])
                    UD2_Phase_temp[i1].append(rows[i1][2])
                
                # 儲存電流資料
                if len(UD2_Current) == 0:
                    UD2_Current += UD2_Current_temp
                
                else:
                    UD2_Current, filled = Merge_data(UD2_Current, UD2_Current_temp)
                    if filled:  # 如果資料有補充
                        data_filled = True
                
                # 儲存相位資料
                if len(UD2_Phase) == 0:
                    UD2_Phase += UD2_Phase_temp
                
                else:
                    UD2_Phase, filled = Merge_data(UD2_Phase, UD2_Phase_temp)
                    if filled:  # 如果資料有補充
                        data_filled = True
            
            # UD7
            elif len(rows[0]) == 2:
                # 儲存電流資料
                if len(UD7_Current) == 0:
                    UD7_Current += rows
                
                else:
                    UD7_Current, filled = Merge_data(UD7_Current, rows)
                    if filled:  # 如果資料有補充
                        data_filled = True
                    
        # 資料確認點            
        #print(UD2_Current)    
        #print(UD2_Phase)
        #print(UD7_Current)
    
    except Exception as e:
        print("合併資料時，發生錯誤", e)

    return UD2_Current, UD2_Phase, UD7_Current, data_filled

# 資料合併模組------------------------------------------------------------------------------------------------

# 聯集合併資料(List A, List B)
def Merge_data(A, B):
    headers = A[0] + B[0][1:]
    body = {}
    data_filled = False

    freq_A = set(row[0] for row in A[1:])
    freq_B = set(row[0] for row in B[1:])

    # 只要 freq 不完全相同，就需要補空白
    if freq_A != freq_B:
        data_filled = True

    # 合併 freq：A、B 都有的 freq，以及獨有的 freq 都要
    all_freqs = sorted(freq_A.union(freq_B))
    default_row = ["" for _ in range(len(headers))]

    # 先處理 A 資料
    for row in A[1:]:
        freq = row[0]
        full_row = default_row[:]
        full_row[:len(row)] = row
        body[freq] = full_row

    # 再處理 B 資料
    for row in B[1:]:
        freq = row[0]
        if freq not in body:
            full_row = default_row[:]
            full_row[0] = freq
            body[freq] = full_row
        body[freq][len(A[0]):] = row[1:]

    # 組合結果
    merged = [headers] + [body[freq] for freq in all_freqs]
    return merged, data_filled

# 掃頻最大點--------------------------------------------------------------------------------------------------

# 找刀具接點 (數據類型, 數據, ΔmA)
def Find_peaks_max(data_type, DATA, ScanErr, show_analysis = None, show_data_labels = None):
    # ScanErr功率關鍵字
    keywords = [f"{i}Vpp" for i in range(200, 401)] + [f"{i} Vpp" for i in range(2, 401)]
    
    # ScanErr放大比率 (預設為 = 1)
    rate = [2 if any(keyword.lower() in text.lower() for keyword in keywords) else 1
            for text in DATA[0]]
    
    # 最大點 / 最小點
    Max_point = ["Max_point"]
    Min_point = ["Min_point"]
    for i2 in range(int(len(DATA[0]) - 1)):
        Max_point.append([])
        Min_point.append([])
        
    # 繪圖控制
    if show_analysis:
        if version == "EE":
            find_peaks_single_drw = False
            find_peaks_all_drw = True
        
        elif version == "CE":
            find_peaks_single_drw = True
            find_peaks_all_drw = False
    
    else:
        find_peaks_single_drw = False
        find_peaks_all_drw = False
    
    try:
        for n in range(1, len(DATA[0])):
            # Step 0：建立資料 ---------------------------------------------------------------------------
            print(f"資料：{DATA[0][n]}")
            Step0_time_start = time.perf_counter()
            
            dStart, dEnd = find_data_ranges(DATA, n) # 取得資料位置
            valleys = []     # 波谷
            prominences = [] # 顯著值
            
            # 建立限制範圍
            Frq_gap = 500 # Hz
            I = dStart + 1
            while DATA[I][n] == "" and I <= dEnd:
                I += 1
            range_limit = int(Frq_gap / (DATA[I][0] - DATA[dStart][0]))
            
            Step0_time_end = time.perf_counter()
            Step0_time = Step0_time_end - Step0_time_start
            #print(f"Step 0：建立資料，處理耗時 = {Step0_time:.10f}")
            
            # Step 1：3點法，找潛在峰值 -------------------------------------------------------------------
            Step1_time_start = time.perf_counter()
            
            gap = find_consecutive_max_length(DATA, n)
            peaks = auto_find_best_gap(find_peak_3point, DATA, n, dStart, dEnd, max_gap = gap)
            peaks_3 = peaks

            Step1_time_end = time.perf_counter()
            Step1_time = Step1_time_end - Step1_time_start
            #print(f"Step 1：3點與5點法，處理耗時 = {Step1_time:.10f}")
                        
            # Step 2：Distance 排除靠近者，保留區間主峰 -----------------------------------------------------
            Step2_time_start = time.perf_counter()
            
            peaks, valleys, prominences = Distance(DATA, n, peaks, valleys, prominences, dStart, dEnd, distance = range_limit)
            peaks_dist = peaks
            
            Step2_time_end = time.perf_counter()
            Step2_time = Step2_time_end - Step2_time_start
            #print(f"Step 2：Distance，處理耗時 = {Step2_time:.10f}")
            
            # Step 3：Prominence 篩選顯著主峰---------------------------------------------------------------
            Step3_time_start = time.perf_counter()
            
            # 搜尋方向調整
            if data_type == "UD2_Phase":
                P_side = "right"
                
            elif data_type == "UD7_Current":
                P_side = "left"
            
            peaks, valleys, prominences = Prominence(DATA, n, peaks, dStart, dEnd, ScanErr, rate, search_range = range_limit, side = P_side)
            peaks_promi = peaks
            valleys_promi = valleys
            #prominences_promi = prominences
            
            Step3_time_end = time.perf_counter()
            Step3_time = Step3_time_end - Step3_time_start
            #print(f"Step 3：Prominence，處理耗時 = {Step3_time:.10f}")
                        
            # Step 4：Ranks 篩選前3名資料 -----------------------------------------------------------------
            Step4_time_start = time.perf_counter()
            
            N = 1
            peaks, valleys, prominences = find_top_n(N, peaks, valleys, prominences)
            
            Step4_time_end = time.perf_counter()
            Step4_time = Step4_time_end - Step4_time_start
            #print(f"Step 4：Ranks，處理耗時 = {Step4_time:.10f}")
            #print()
            print(f"找共振點，總共耗時 = {Step0_time + Step1_time + Step2_time + Step3_time + Step4_time:.10f}")
            print()
            print("---------------")
            print()
            
            # Step 5：儲存 Max / Min point ---------------------------------------------------------------
            Max_point[n] = " / ".join(f"{DATA[p][0]}:{DATA[p][n]}" for p in peaks)
            Min_point[n] = " / ".join(f"{DATA[p][0]}:{DATA[p][n]}" for p in valleys)
            
            # 繪圖分析 -----------------------------------------------------------------------------------
            # 獨立圖表
            if find_peaks_single_drw:
                plot_final_peak(data_type, DATA, n, dStart, dEnd, peaks, valleys, drw_show_data_labels = show_data_labels)
            
            # 完整圖表
            if find_peaks_all_drw:
                plot_find_max_steps(data_type, DATA, n, dStart, dEnd,
                                    peaks_3,
                                    peaks_dist,
                                    peaks_promi, valleys_promi,
                                    peaks, valleys,
                                    drw_show_data_labels = show_data_labels
                                    )
            
        # 確認資料
        print(f"{data_type} 掃頻最大點：")
        print(DATA[0])
        print(Max_point)
        print(Min_point)
            
    except Exception as e:
        print("⚠️找共振點發生錯誤：", e)
        messagebox.showerror("錯誤", "找掃頻最大點時，發生錯誤！")
    
    return Max_point, Min_point

# 找出有資料的起始和終止位置(資料, 資料位置)
def find_data_ranges(DATA, column_index):
    Start, End = None, None
    for i, row in enumerate(DATA[1:]):
        value = row[column_index]
        if value != "":  # 如果有資料
            if Start is None:
                Start = i  # 記錄起始位置
            End = i        # 更新終止位置
    Start += 1
    End   += 1
    return Start, End

# 計算數據最大重複值(平台)
def find_consecutive_max_length(DATA, n):
    max_len = 0
    count = 0
    prev_value = None

    for i in range(len(DATA)):
        current = DATA[i][n]
        if current == "":
            continue  # 空值不計算也不重設

        if current == prev_value:
            count += 1
        else:
            count = 1
            prev_value = current

        max_len = max(max_len, count)

    return max_len

# 自動找最佳間距
def auto_find_best_gap(method_func, DATA, n, dStart, dEnd, max_gap = 5):
    get_peaks = set()
    
    for gap in range(1, max_gap + 1):
        peaks = method_func(DATA, n, dStart, dEnd, gap)
        get_peaks = get_peaks | set(peaks)

    get_peaks = list(get_peaks)        
    get_peaks.sort()
    
    return get_peaks

# 3點找PEAK
def find_peak_3point(DATA, n, dStart, dEnd, gap):
    peaks_index = []

    for i in range(dStart, dEnd):
        if DATA[i][n] == "":
            continue

        # 往左找 gap 個有效資料點
        left = i - 1
        left_count = 0
        while left >= dStart:
            if DATA[left][n] != "":
                left_count += 1
                if left_count == gap:
                    break
            left -= 1
        if left < dStart or DATA[left][n] == "":
            continue  # 左邊不足 gap 個有效點

        # 往右找 gap 個有效資料點
        right = i + 1
        right_count = 0
        while right < dEnd:
            if DATA[right][n] != "":
                right_count += 1
                if right_count == gap:
                    break
            right += 1
        if right >= dEnd or DATA[right][n] == "":
            continue  # 右邊不足 gap 個有效點

        # 比較大小
        if DATA[i][n] > DATA[left][n] and DATA[i][n] > DATA[right][n]:
            peaks_index.append(i)

    return peaks_index

# Distance 排除靠近者，保留區間主峰
def Distance(DATA, n, peaks, valleys, prominences, dStart, dEnd, distance = 1):
    # 將 peak 座標排序（保險）
    peaks = sorted(peaks)

    # 預先建立值陣列（避免重複存取 DATA）
    peak_values = {p: DATA[p][n] for p in peaks if isinstance(DATA[p][n], (int, float))}

    # 標記哪些 peak 要保留（預設全部保留）
    keep_flags = {p: True for p in peaks}

    # 主迴圈：滑動窗口 + 比較距離
    for i in range(len(peaks)):
        if not keep_flags[peaks[i]]:
            continue  # 若已經被淘汰，就跳過

        group  = [peaks[i]]
        values = [peak_values.get(peaks[i], -float('inf'))]

        for j in range(i + 1, len(peaks)):
            if abs(peaks[j] - peaks[i]) <= distance:
                if keep_flags[peaks[j]]:
                    group.append(peaks[j])
                    values.append(peak_values.get(peaks[j], -float('inf')))
            else:
                break  # 超過範圍就不比了（因為 peaks 是排序的）

        if len(group) > 1:
            # 找出該區域最大者
            ranked = ranks(values)
            winner_index = ranked.index(1)
            winner = group[winner_index]

            for k, p in enumerate(group):
                if p != winner:
                    keep_flags[p] = False

    # 根據 keep_flags 重建 peaks / valleys / prominences
    new_peaks       = []
    new_valleys     = []
    new_prominences = []

    for i, p in enumerate(peaks):
        if keep_flags[p]:
            new_peaks.append(p)
            if i < len(valleys):
                new_valleys.append(valleys[i])
            if i < len(prominences):
                new_prominences.append(prominences[i])

    return new_peaks, new_valleys, new_prominences

# Prominence
def Prominence(DATA, n, peaks, dStart, dEnd, prominence, rate, search_range = None, side = 'both'):
    new_peaks = []
    valleys = []
    prominences = []

    for peak_index in peaks:
        peak_value = DATA[peak_index][n] # 峰值
        left_valley_value  = peak_value  # 左波谷值
        right_valley_value = peak_value  # 右波谷值
        left_valley_index  = peak_index  # 左波谷值序
        right_valley_index = peak_index  # 右波谷值序

        # 搜尋左谷底
        if side in ('left', 'both'):
            left = peak_index
            step_left = 0
            while left > dStart and (search_range is None or step_left < search_range):
                left -= 1

                if isinstance(DATA[left][n], (int, float)):
                    step_left += 1
                    
                    if DATA[left][n] > peak_value:
                        break
                    if DATA[left][n] < left_valley_value:
                        left_valley_index = left
                        left_valley_value = DATA[left][n]

        # 搜尋右谷底
        if side in ('right', 'both'):
            right = peak_index
            step_right = 0
            while right < dEnd and (search_range is None or step_right < search_range):
                right += 1

                if isinstance(DATA[right][n], (int, float)):
                    step_right += 1
                    
                    if DATA[right][n] > peak_value:
                        break
                    if DATA[right][n] < right_valley_value:
                        right_valley_index = right
                        right_valley_value = DATA[right][n]

        # 計算 prominence 值
        if side == 'left':
            prominence_value = peak_value - left_valley_value
            chosen_valley = left_valley_index
            
        elif side == 'right':
            prominence_value = peak_value - right_valley_value
            chosen_valley = right_valley_index
            
        else:  # both
            if left_valley_value >= right_valley_value:
                chosen_valley = left_valley_index
                
            elif left_valley_value < right_valley_value:
                chosen_valley = right_valley_index
                
            prominence_value = peak_value - max(left_valley_value, right_valley_value)

        if prominence_value >= prominence * rate[n]:
            new_peaks.append(peak_index)
            valleys.append(chosen_valley)
            prominences.append(prominence_value)

    return new_peaks, valleys, prominences

# 找出前 N 名
def find_top_n(N, peaks, valleys, prominences):
    if len(peaks) > N:
        prominence_ranks = ranks(prominences)
        # 過濾出排名在 N 以內的元素
        selected_indices = [i for i, r in enumerate(prominence_ranks) if r <= N]
        
        # 取出篩選後的新資料
        peaks = [peaks[i] for i in selected_indices]
        valleys = [valleys[i] for i in selected_indices]
        prominences = [prominences[i] for i in selected_indices]
    
    return peaks, valleys, prominences

# 電流排名篩選
def ranks(Current):
    # 把原始資料配上索引
    indexed = list(enumerate(Current))

    # 根據數值從大到小排序
    sorted_indexed = sorted(indexed, key = lambda x: x[1], reverse = True)

    # 建立結果 list，初始全 0
    ranks = [0] * len(Current)

    # 記錄每個原始位置的排名
    for rank, (i, val) in enumerate(sorted_indexed, start = 1):
        ranks[i] = rank
    
    return ranks

# Find Peaks analysis ---------------------------------------------------------------------------------------
# 只顯示最終共振點結果
def plot_final_peak(data_type, DATA, channel_index, dStart, dEnd, peaks_final, valleys_final, drw_show_data_labels = None):
    # 整理 x 和 y 資料
    x = [row[0] for row in DATA[dStart:dEnd] if row[channel_index] not in [None, ""]]
    y = [row[channel_index] for row in DATA[dStart:dEnd] if row[channel_index] not in [None, ""]]

    # 創建圖表
    fig, ax = plt.subplots(figsize = (14, 10))

    # 繪製原始數據曲線
    ax.plot(x, y, label = 'Original Curve', color = 'lightgray')

    # 標註最終的 Peaks 和 Valleys
    ax.scatter([DATA[p][0] for p in peaks_final], [DATA[p][channel_index] for p in peaks_final],
               color = 'green', marker = '*', s = 120, label = 'Peaks (Final)')
    ax.scatter([DATA[v][0] for v in valleys_final], [DATA[v][channel_index] for v in valleys_final],
               color = 'gray', marker = 'x', label = 'Valleys (Final)')

    # 繪製峰值和谷值之間的垂直線
    for i, peak in enumerate(peaks_final):
        valley = valleys_final[i]
        ax.vlines(DATA[peak][0], DATA[valley][channel_index], DATA[peak][channel_index],
                  color = 'black', linestyle = '--', alpha = 0.5)
        
    # 顯示資料標籤
    if drw_show_data_labels:
        # 取得目前 y 軸範圍（原始資料範圍）
        y_min, y_max = ax.get_ylim()
        peak_y_offset   = (y_max - y_min) * 0.02
        valley_y_offset = (y_max - y_min) * 0.035
    
        # 放標籤，順便記錄標籤最高和最低 y 座標
        label_ys = []
        for i, peak in enumerate(peaks_final):
            valley = valleys_final[i]
        
            ax.vlines(DATA[peak][0], DATA[valley][channel_index], DATA[peak][channel_index],
                              color = 'black', linestyle = '--', alpha = 0.5)
        
            peak_y = DATA[peak][channel_index] + peak_y_offset
            ax.text(DATA[peak][0], peak_y,
                            f'{DATA[peak][0]}, {DATA[peak][channel_index]}',
                            color = 'black', fontsize = 12, ha = 'center', clip_on = True)
            label_ys.append(peak_y)
        
            valley_y = DATA[valley][channel_index] - valley_y_offset
            ax.text(DATA[valley][0], valley_y,
                            f'{DATA[valley][0]}, {DATA[valley][channel_index]}',
                            color = 'black', fontsize = 12, ha = 'center', clip_on = True)
            label_ys.append(valley_y)
        
        if label_ys:
            max_label_y = max(label_ys)
            min_label_y = min(label_ys)
            new_y_min   = min(y_min, min_label_y - valley_y_offset * 1.5)
            new_y_max   = max(y_max, max_label_y + peak_y_offset * 2)
            ax.set_ylim(new_y_min, new_y_max)
        
        else:
            ax.set_ylim(y_min, y_max)    
        
    # 保留原圖表標題
    ax.set_title('Peak & Valley final result')

    # 設置獨立標題
    fig.suptitle(DATA[0][channel_index], fontsize = 12, fontweight = 'bold')
    
    # 設置 x 軸標籤
    ax.set_xlabel('Frequency [Hz]')

    # 根據 data_type 設置 y 軸標籤
    if data_type == "UD2_Phase":
        ax.set_ylabel('Phase [degree]')
        
    elif data_type == "UD7_Current":
        ax.set_ylabel('Current [mA]')

    # 顯示網格
    ax.grid(True)

    # 顯示圖例
    ax.legend()

    # 自動調整布局
    plt.tight_layout()
    plt.get_current_fig_manager().window.showMaximized()
    plt.show()

# 3點法 + Distance + Prominence 圖表
def plot_find_max_steps(
                        data_type, DATA, channel_index, dStart, dEnd,
                        peaks_3, 
                        peaks_distance,
                        peaks_promi, valleys_promi,
                        peaks_final, valleys_final,
                        drw_show_data_labels = None
                        ):
    
    x = [row[0] for row in DATA[dStart:dEnd] if row[channel_index] not in [None, ""]]
    y = [row[channel_index] for row in DATA[dStart:dEnd] if row[channel_index] not in [None, ""]]

    fig, axes = plt.subplots(2, 2, figsize = (14, 10), sharex = True, sharey = True)
    fig.suptitle(DATA[0][channel_index], fontsize = 12, fontweight = 'bold')

    # Step 1: 3點法找出的 peaks
    axes[0, 0].plot(x, y, label = 'Original Curve', color = 'lightgray')
    axes[0, 0].scatter([DATA[p][0] for p in peaks_3], [DATA[p][channel_index] for p in peaks_3],
                       color = 'red', marker = 'o', label = 'Peaks (by 3-point)')
    axes[0, 0].set_title('Step 1: 3-Point Filtered Peaks', fontweight = 'bold')
    axes[0, 0].legend()

    # Step 2: Distance 篩選後保留的 peaks
    axes[0, 1].plot(x, y, label = 'Original Curve', color = 'lightgray')
    axes[0, 1].scatter([DATA[p][0] for p in peaks_distance], [DATA[p][channel_index] for p in peaks_distance],
                       color = 'blue', marker = 's', label = 'Peaks (by Distance)')
    axes[0, 1].set_title('Step 2: Distance-Filtered Peaks', fontweight='bold')
    axes[0, 1].legend()

    # Step 3: Prominence 結果
    axes[1, 0].plot(x, y, label = 'Original Curve', color = 'lightgray')
    axes[1, 0].scatter([DATA[p][0] for p in peaks_promi], [DATA[p][channel_index] for p in peaks_promi],
                       color = 'green', marker = '^', label = 'Peaks (by Prominence)')
    axes[1, 0].scatter([DATA[v][0] for v in valleys_promi], [DATA[v][channel_index] for v in valleys_promi],
                       color = 'gray', marker = 'x', label = 'Valleys (by Prominence)')
    for i, peak in enumerate(peaks_promi):
        valley = valleys_promi[i]
        axes[1, 0].vlines(DATA[peak][0], DATA[valley][channel_index], DATA[peak][channel_index],
                          color = 'black', linestyle = '--', alpha = 0.5)
    axes[1, 0].set_title('Step 3: Prominence-Filter Peaks', fontweight = 'bold')
    axes[1, 0].legend()
    
    # Step 4: Ranks 結果
    axes[1, 1].plot(x, y, label = 'Original Curve', color = 'lightgray')
    axes[1, 1].scatter([DATA[p][0] for p in peaks_final], [DATA[p][channel_index] for p in peaks_final],
                    color = 'gold', marker = '*', s = 120, label = 'Peak (Final)')
    axes[1, 1].scatter([DATA[v][0] for v in valleys_final], [DATA[v][channel_index] for v in valleys_final],
                       color = 'gray', marker = 'x', label = 'Valley (Final)')
    for i, peak in enumerate(peaks_final):
        valley = valleys_final[i]
        axes[1, 1].vlines(DATA[peak][0], DATA[valley][channel_index], DATA[peak][channel_index],
                          color = 'black', linestyle = '--', alpha = 0.5)
    axes[1, 1].set_title('Step 4: Top 1 Peak', fontweight = 'bold')
    axes[1, 1].legend()
    
    # 顯示資料標籤
    if drw_show_data_labels:
        # 取得目前 y 軸範圍（原始資料範圍）
        y_min, y_max    = axes[1, 1].get_ylim()
        peak_y_offset   = (y_max - y_min) * 0.035
        valley_y_offset = (y_max - y_min) * 0.065
    
        # 放標籤，順便記錄標籤最高和最低 y 座標
        label_ys = []
        for i, peak in enumerate(peaks_final):
            valley = valleys_final[i]
        
            axes[1, 1].vlines(DATA[peak][0], DATA[valley][channel_index], DATA[peak][channel_index],
                              color = 'black', linestyle = '--', alpha = 0.5)
        
            peak_y = DATA[peak][channel_index] + peak_y_offset
            axes[1, 1].text(DATA[peak][0], peak_y,
                            f'{DATA[peak][0]}, {DATA[peak][channel_index]}',
                            color = 'black', fontsize = 10, ha = 'center', clip_on = True)
            label_ys.append(peak_y)
        
            valley_y = DATA[valley][channel_index] - valley_y_offset
            axes[1, 1].text(DATA[valley][0], valley_y,
                            f'{DATA[valley][0]}, {DATA[valley][channel_index]}',
                            color = 'black', fontsize = 10, ha = 'center', clip_on = True)
            label_ys.append(valley_y)
        
        if label_ys:
            max_label_y = max(label_ys)
            min_label_y = min(label_ys)
            new_y_min   = min(y_min, min_label_y - valley_y_offset * 1.5)
            new_y_max   = max(y_max, max_label_y + peak_y_offset * 2)
            axes[1, 1].set_ylim(new_y_min, new_y_max)
        
        else:
            axes[1, 1].set_ylim(y_min, y_max)
    
    # 坐標軸標題
    for i in range(2):
        for j in range(2):
            ax = axes[i, j]
            if data_type == "UD2_Phase":
                ax.set_ylabel('Phase [degree]')
                
            elif data_type == "UD7_Current":
                ax.set_ylabel('Current [mA]')
            
            if i == 1:
                ax.set_xlabel('Frequency [Hz]') 
            
            ax.grid(True)

    plt.tight_layout()
    plt.get_current_fig_manager().window.showMaximized()
    plt.show()

# 繪圖模組----------------------------------------------------------------------------------------------------

xMin = lambda x: (x // 1000) * 1000 if (x % 1000) > 0 else x # X軸最小值調整
xMax = lambda x: ((x // 1000) + 1) * 1000 if (x % 1000) > 0 else x # X軸最大值調整

# 圖表標題設定 (圖表, 1 pt = 100 size)
def set_chart_title_size(chart, size = 1400):
    paraprops = ParagraphProperties()
    paraprops.defRPr = CharacterProperties(sz = size)

    for para in chart.title.tx.rich.paragraphs:
        para.pPr = paraprops

# Excel 圖表繪製 (標題, 單位, 數據, 顏色, 線型, 線寬, Excel分頁, 資料分割段)
def Drawing(title, unit, DATA, Lcolors, Ltypes, Lwidth, ws, R):
    # XY散佈圖
    chart = ScatterChart()
    chart.style = 13
    chart.title = title
    set_chart_title_size(chart, size = 1400)
    chart.x_axis.title = "Frequency"

    if title != "Current & Phase":
        chart.y_axis.title = unit

    else:
        # 左Y軸
        chart.y_axis.title = "Current [mA]"
        
        # 右Y軸
        chart2 = ScatterChart()
        chart2.y_axis.title = "Phase [Degree]"
        chart2.y_axis.axId = 200
        chart2.y_axis.crosses = 'max'
        chart2.y_axis.majorGridlines = None
        chart2.y_axis.majorTickMark = 'out'
    
    # X軸
    xvalues = Reference(ws, min_col = 1, min_row = 2, max_row = len(DATA))
    chart.x_axis.scaling.min = xMin(DATA[1][0])   # Minimum value for x-axis
    chart.x_axis.scaling.max = xMax(DATA[-1][0])  # Maximum value for x-axis
    
    # X軸刻度間距
    if (xMax(DATA[-1][0]) - xMin(DATA[1][0]))   <= 3000:
        chart.x_axis.majorUnit = 500
        
    elif (xMax(DATA[-1][0]) - xMin(DATA[1][0])) <= 6000:
        chart.x_axis.majorUnit = 1000
        
    elif (xMax(DATA[-1][0]) - xMin(DATA[1][0])) <= 12000:
        chart.x_axis.majorUnit = 2000
        
    else:
        chart.x_axis.majorUnit = 5000

    # Y軸
    for i in range(len(R)):
        for y in range(R[i][0], R[i][1]):
            yvalues = Reference(ws, min_col = y, min_row = 1, max_row = len(DATA))
            series = Series(yvalues, xvalues, title_from_data = True)
            
            # 設定線寬、顏色、線型
            line_properties = LineProperties(w = Lwidth, solidFill = Lcolors[y - R[i][0]], prstDash = Ltypes[y - R[i][0]])
            series.graphicalProperties.line = line_properties
            
            # 資料存檔
            if i == 0:
                chart.series.append(series)
                
            elif i == 1:
                chart2.series.append(series)
    
    # 設定線的樣式平滑曲線
    for x in range(R[0][2]):
        chart.series[x].smooth = True
        
    if title == "Current & Phase":
        for x in range(R[1][2]):
            chart2.series[x].smooth = True
            
    if title == "Current & Phase":
        chart += chart2
    
    # 圖表儲存位置
    adress = Drawing_adress(len(DATA[0])) + str("1")
    
    # 圖表大小
    chart.height = 7.5  # 設置高度
    chart.width = 17   # 設置寬度
    return chart, adress, xMax(DATA[-1][0]), xMin(DATA[1][0])

# 建立Excel圖表儲存位置
def Drawing_adress(n):
    adress = []
    n += 1
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        adress.append(chr(remainder + ord('A')))
    return ''.join(reversed(adress))

# 隱藏的空格與空值補線
def set_excel_chart_options(file_path):
    excel = win32.DispatchEx("Excel.Application")  # 使用 DispatchEx 確保背景執行
    excel.Visible = False                          # 設定為不可見
    excel.DisplayAlerts = False                    # 關閉提示

    try:
        # 開啟文件
        workbook = excel.Workbooks.Open(file_path)
        
        # 取得所有分頁
        for sheet in workbook.Sheets:
            # 迭代該分頁中的所有圖表
            for chart_object in sheet.ChartObjects():
                chart = chart_object.Chart
                # 設定「連接資料點的線」
                chart.DisplayBlanksAs = 3  # 3 表示連接空白資料點 (xlInterpolated)

        # 儲存
        workbook.Save()

    finally:
        # 確保清理資源並關閉 Excel
        workbook.Close(SaveChanges = True)
        excel.Quit()

# GUI介面----------------------------------------------------------------------------------------------------

class UD_App:
    def __init__(self, root):
        self.root = root
        self.root.title("UD2/UD7 merge files")
        self.window_width  = 260
        self.window_height = 430
        self.root.geometry(f"{self.window_width}x{self.window_height}")  # 設置窗口大小
        self.root.minsize(self.window_width, self.window_height) # 限制視窗大小
        self.root.maxsize(self.window_width, self.window_height) # 限制視窗大小
        self.root.resizable(False, False)

        # 瀏覽資料夾框架
        self.f_path_frame = ttk.LabelFrame(root, text = "檔案路徑：", relief = "groove", borderwidth = 2)
        self.f_path_frame.place(x = 15, y = 10, width = 230, height = 60)
        self.f_path_frame.config(style = "Dashed.TFrame")
        
        # 創建Listbox
        self.listbox = tk.Listbox(self.f_path_frame, width = 23, height = 1)
        self.listbox.grid(row = 0, column = 0, padx = 5, pady = 5)

        # 創建瀏覽資料夾按鈕
        self.browse_button = ttk.Button(self.f_path_frame, text = "瀏覽...", command = self.browse_folder, width = 6.5)
        self.browse_button.grid(row = 0, column = 1, padx = 0, pady = 5)

        # 儲存最新選擇的資料夾路徑
        self.latest_folder = os.getcwd()
        self.update_listbox()

        # 第一行的選項 --------------------------------------------------------------------------------------
        self.frame1 = ttk.LabelFrame(root, text = "UD驅動器數據資料：", relief = "groove", borderwidth = 2)
        self.frame1.place(x = 15, y = 80, width = 230, height = 55)
        self.frame1.config(style = "Dashed.TFrame")

        self.var_c = tk.BooleanVar(value = True) # c = 電流
        self.var_p = tk.BooleanVar(value = True) # p = 相位

        self.check_c = ttk.Checkbutton(self.frame1, text = "電流 (Current)", variable = self.var_c)
        self.check_p = ttk.Checkbutton(self.frame1, text = "相位 (Phase)",   variable = self.var_p)

        self.check_c.grid(row = 0, column = 0, padx = 5, pady = 5)
        self.check_p.grid(row = 0, column = 1, padx = 5, pady = 5)

        
        # 掃頻設定 ---------------------------------------------------------------------------------------------
        self.scan_frame = ttk.LabelFrame(root, text = "           ", relief = "groove", borderwidth = 2)
        self.scan_frame.place(x = 15, y = 145, width = 230, height = 115)
        self.scan_frame.config(style = "Dashed.TFrame")
        
        self.var_max = tk.BooleanVar(value = True) # max_point = 找刀把共振點
        self.check_max = ttk.Checkbutton(self.root, text = "掃頻最大點：", variable = self.var_max, command = self.scan_set_state)
        self.check_max.place(x = 22, y = 142)
        
        # UD2 掃頻設定
        self.UD2_scan_label = ttk.Label(self.scan_frame, text = "UD2相位差：")
        self.UD2_scan_unit  = ttk.Label(self.scan_frame, text = "°")
        
        self.UD2_scan_combo_var = tk.StringVar()
        self.UD2_scan_combo = ttk.Combobox(self.scan_frame, 
                                  textvariable = self.UD2_scan_combo_var, 
                                  values = [str(i) for i in range(2, 16)], 
                                  width  = 3,
                                  state  = 'readonly')
        self.UD2_scan_combo.current(4)  # 預設 = 6
        
        self.UD2_scan_label.grid(row = 1, column = 0, padx = 5, pady = 5, sticky = "w")
        self.UD2_scan_combo.grid(row = 1, column = 1, padx = 0, pady = 5)
        self.UD2_scan_unit.grid (row = 1, column = 2, padx = 0, pady = 5)
        
        # UD7 掃頻設定
        self.UD7_scan_label = ttk.Label(self.scan_frame, text = "UD7電流差：")
        self.UD7_scan_unit  = ttk.Label(self.scan_frame, text = " mA")
        
        self.UD7_scan_combo_var = tk.StringVar()
        self.UD7_scan_combo = ttk.Combobox(self.scan_frame, 
                                  textvariable = self.UD7_scan_combo_var, 
                                  values = [str(i) for i in range(10, 251)], 
                                  width  = 3,
                                  state  = 'readonly')
        self.UD7_scan_combo.current(10)  # 預設 = 20

        self.UD7_scan_label.grid(row = 2, column = 0, padx = 5, pady = 5, sticky = "w")
        self.UD7_scan_combo.grid(row = 2, column = 1, padx = 0, pady = 5)
        self.UD7_scan_unit.grid (row = 2, column = 2, padx = 0, pady = 5)
        
        # 分析圖表
        self.peak_drw_var = tk.BooleanVar()
        self.check_peak_drw = ttk.Checkbutton(self.scan_frame, text = "共振點分析圖表", command = self.data_labels_state, variable = self.peak_drw_var)
        self.check_peak_drw.grid(row = 3, column = 0, padx = 5, pady = 5, columnspan = 2, sticky = "w")
        
        self.data_labels_var = tk.BooleanVar()
        self.data_labels = ttk.Checkbutton(self.root, text = "顯示數據資料", variable = self.data_labels_var, state = "disabled")
        self.data_labels.place(x = 135, y = 229)

        # 下拉選單 ---------------------------------------------------------------------------------------------
        self.frame2 = ttk.LabelFrame(root, text = "請選擇數量：", relief = "groove", borderwidth = 2)
        self.frame2.place(x = 15, y = 270, width = 125, height = 115)
        self.frame2.config(style = "Dashed.TFrame")

        # 刀把數量
        self.label1 = ttk.Label(self.frame2, text = "刀把數量")
        self.label1.grid(row = 0, column = 0, padx = 5, pady = 5)

        self.combo_var1 = tk.StringVar()
        self.combo = ttk.Combobox(self.frame2, 
                                  textvariable = self.combo_var1, 
                                  values = [str(i) for i in range(1, 1000)], 
                                  width = 3,
                                  state = 'readonly')
        self.combo.grid(row = 0, column = 1, padx = 5, pady = 5)
        self.combo.current(0)  # 預設選擇第一個選項

        # 刀具數量
        self.label2 = ttk.Label(self.frame2, text = "刀具數量")
        self.label2.grid(row = 1, column = 0, padx = 5, pady = 5)

        self.combo_var2 = tk.StringVar()
        self.combo2 = ttk.Combobox(self.frame2, 
                                   textvariable = self.combo_var2, 
                                   values = [str(i) for i in range(0, int(len(linetypes) + 1))], 
                                   width = 3,
                                   state = 'readonly')
        self.combo2.grid(row = 1, column = 1, padx = 5, pady = 5)
        self.combo2.current(0)  # 預設選擇第一個選項
        
        # 循環次數
        self.label4 = ttk.Label(self.frame2, text = "循環合併")
        self.label4.grid(row = 2, column = 0, padx = 5, pady = 5)

        self.combo_var4 = tk.StringVar()
        self.combo4 = ttk.Combobox(self.frame2,
                                   textvariable = self.combo_var4, 
                                   values = [i for i in range(1, 101)], 
                                   width = 3,
                                   state = 'readonly')
        self.combo4.grid(row = 2, column = 1, padx = 5, pady = 5)
        self.combo4.current(0)  # 預設選擇第一個選項

        # 進階選單 ------------------------------------------------------------------------------------------
        self.frame3 = ttk.LabelFrame(root, text = "進階選項：", relief = "groove", borderwidth = 2)
        self.frame3.place(x = 145, y = 270, width = 100, height = 115)
        self.frame3.config(style = "Dashed.TFrame")

        self.var_color    = tk.BooleanVar()
        self.var_linetype = tk.BooleanVar()
        self.combo_var3   = tk.StringVar() # 線寬選單
        
        # 線寬清單
        combo3_values = []  # 線寬空集合
        Lw = 0.5            # 起始
        for i in range(10):
            combo3_values.append("線寬:" + str(Lw) + " pt")
            Lw += 0.5 # 線寬間距
            
        self.combo3 = ttk.Combobox(self.frame3, 
                                   textvariable = self.combo_var3, 
                                   values = combo3_values, 
                                   width = 9,
                                   state = 'readonly')
        self.combo3.current(1)

        self.check_color    = ttk.Checkbutton(self.frame3, text = "單色線條", variable = self.var_color)
        self.check_linetype = ttk.Checkbutton(self.frame3, text = "實線線條", variable = self.var_linetype)

        self.combo3.grid        (row = 0, column = 1, padx = 5, pady = 5, sticky = "w")
        self.check_color.grid   (row = 1, column = 0, padx = 5, pady = 5, columnspan = 2, sticky = "w")
        self.check_linetype.grid(row = 2, column = 0, padx = 5, pady = 5, columnspan = 2, sticky = "w")
        
        # 按鈕框架 -----------------------------------------------------------------------------------------------
        self.button_frame = ttk.Frame(root)
        self.button_frame.place(x = 40, y = 390)

        # 創建按鈕
        self.run_button   = ttk.Button(self.button_frame, text = "執行", command = self.run_action,   width = 10)
        self.close_button = ttk.Button(self.button_frame, text = "離開", command = self.close_action, width = 10)

        self.run_button.grid  (row = 0, column = 0, padx = 5, pady = 5)
        self.close_button.grid(row = 0, column = 1, padx = 5, pady = 5)

    # 瀏覽資料夾
    def browse_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.latest_folder = folder_path
            self.update_listbox()
            self.update_checkbutton_state()
            
    # 清除Listbox的內容，並插入最新選擇的資料夾路徑    
    def update_listbox(self):
        self.listbox.delete(0, tk.END)
        self.listbox.insert(tk.END, self.latest_folder)

    def update_checkbutton_state(self):
        self.file_count = int(len(get_files_in_dir(self.latest_folder)))

        # 創建一個包含從 1 到 file_count 的列表
        values = list(range(1, self.file_count + 1))
        
        # 更新刀把數量的值
        self.combo['values'] = values
        if values:
            self.combo.set(values[-1])

        # 更新刀具數量的值        
        self.combo2.current(0)
    
    # 掃頻設定狀態
    def scan_set_state(self):
        if self.var_max.get():
            self.UD2_scan_combo.config(state = "readonly")
            self.UD7_scan_combo.config(state = "readonly")
            self.check_peak_drw.config(state = "readonly")
        
        else:
            self.UD2_scan_combo.config(state = "disabled")
            self.UD7_scan_combo.config(state = "disabled")
            self.check_peak_drw.config(state = "disabled")
            self.data_labels.config   (state = "disabled")
            self.peak_drw_var.set   (False)
            self.data_labels_var.set(False)
    
    # 數據資料狀態
    def data_labels_state(self):
        if self.peak_drw_var.get():
            self.data_labels.config(state = "readonly")
        
        else:
            self.data_labels_var.set(False)
            self.data_labels.config(state = "disabled")
    
    # 執行合併
    def run_action(self):
        # 取得合併的資料
        UD2_Current, UD2_Phase, UD7_Current, data_filled = UD_data(self.latest_folder)
        
        # 合併UD2和UD7的資料
        if len(UD2_Current) > 0 and len(UD7_Current) > 0:
            print("提示：UD2與UD7資料合併模式！")
            Current, data_filled = Merge_data(UD2_Current, UD7_Current)
            d_type = "UD2+UD7" # 改變檔名
            
        elif len(UD2_Current) > 0 and len(UD7_Current) == 0:
            print("提示：UD2資料合併模式！")
            Current = UD2_Current
            d_type = "UD2" # 改變檔名

        elif len(UD2_Current) == 0 and len(UD7_Current) > 0:
            print("提示：UD7資料合併模式！")
            Current = UD7_Current
            d_type = "UD7" # 改變檔名
        
        Phase = UD2_Phase
        
        # 刀把數量
        USTH_number = int(self.combo_var1.get())
        
        # 刀具數量
        tool_number = int(self.combo_var2.get())
        
        # 循環次數
        cycle = int(self.combo_var4.get())

        Ncolors = [] # 增量預設顏色
        Lcolors = [] # 增量顏色
        Ltypes  = [] # 增量線型
        
        # 增量預設顏色
        if USTH_number > len(colors):
            Ncolors = (USTH_number // len(colors)) * colors
            for i1 in range(USTH_number % len(colors)):
               Ncolors.append(colors[i1])
        else:
            Ncolors = colors
        
        # 增量顏色
        for i2 in range(USTH_number):
            if tool_number <= 1:
                tool_number = 1
                n = 1
            else:
                n = tool_number
                
            # 圖表顏色、線型增量
            for j1 in range(n):
                Lcolors.append(Ncolors[i2])
                Ltypes.append(linetypes[j1])
        
        # 循環合併
        Lcolors = Lcolors * cycle
        Ltypes  = Ltypes  * cycle
        
        # 圖表統一顏色
        if self.var_color.get():
            for i3 in range(len(Lcolors)):
                Lcolors[i3] = Ncolors[0]
                
        # 圖表統一線型
        if self.var_linetype.get():
            for i4 in range(len(Ltypes)):
                Ltypes[i4] = linetypes[0]

        # 線寬
        Lwidth = float(self.combo_var3.get()[3:6])
        Lwidth = int(Lwidth * 12700.2)

        if USTH_number * tool_number * cycle >= self.file_count and (UD2_Current != [] or UD2_Phase != [] or UD7_Current != []):
            try:
                # 建立新工作表
                wb = Workbook()
                
                # 將資料存進Current的分頁
                if self.var_c.get():
                    # 建立分頁
                    ws = wb.active
                    ws.title = "Current"
                    
                    # 儲存點資料
                    for row in Current:
                        ws.append(row)
                    
                    # 儲存掃頻最大點
                    if self.var_max.get() and len(UD7_Current) > 0:
                        UD7_max_point, UD7_min_point = Find_peaks_max("UD7_Current", 
                                                                      UD7_Current, 
                                                                      int(self.UD7_scan_combo_var.get()), # 只適用UD7
                                                                      show_analysis = self.peak_drw_var.get(),
                                                                      show_data_labels = self.data_labels_var.get())
                        
                        # 跳過UD2非判斷點位置
                        if len(UD2_Current) > 0:
                            for i5 in range(len(UD2_Current[0]) - 1):
                                UD7_max_point.insert(1, "")
                                UD7_min_point.insert(1, "")
                                
                        ws.append(UD7_max_point)
                        ws.append(UD7_min_point)
                    
                    # 儲存圖表
                    if len(Current[0]) <= 256:
                        R = [ [2, len(Current[0]) + 1, len(Current[0]) - 1] ] # 起始, 終止, 平滑次數
                        chart, adress, xMax, xMin = Drawing("Current", "Current [mA]", Current, Lcolors, Ltypes, Lwidth, ws, R)
                        ws.add_chart(chart, adress)
                    else:
                        messagebox.showwarning("圖表錯誤", "繪製圖表發生錯誤：\n\nExcel一張圖表最多只能容納 255 個資料數列！")
                
                # 將資料存進Phase的分頁
                if self.var_p.get() and len(UD2_Phase) > 0:
                    # 建立分頁
                    if self.var_c.get() == 0: # 判斷Current無勾選
                        ws = wb.active
                        ws.title = "Phase"
                        
                    else:
                        ws = wb.create_sheet(title = "Phase")
                    
                    # 儲存點資料
                    for row in Phase:
                        ws.append(row)
                    
                    # 儲存掃頻最大點
                    if self.var_max.get() and len(UD2_Phase) > 0:
                        UD2_max_point, UD2_min_point = Find_peaks_max("UD2_Phase", 
                                                                      UD2_Phase, 
                                                                      int(self.UD2_scan_combo_var.get()), # 只適用UD2
                                                                      show_analysis = self.peak_drw_var.get(),
                                                                      show_data_labels = self.data_labels_var.get())
                        ws.append(UD2_max_point)
                        ws.append(UD2_min_point)
                    
                    # 儲存圖表
                    if len(Phase[0]) <= 256:
                        R = [ [2, len(Phase[0]) + 1, len(Phase[0]) - 1] ] # 起始, 終止, 平滑次數  
                        chart, adress, xMax, xMin = Drawing("Phase", "Phase [degree]", Phase, Lcolors, Ltypes, Lwidth, ws, R)
                        ws.add_chart(chart, adress)
                    else:
                        messagebox.showwarning("圖表錯誤", "繪製圖表發生錯誤：\n\nExcel一張圖表最多只能容納 255 個資料數列！")
                
                # 只適用UD2
                if self.var_c.get() and self.var_p.get() and len(UD2_Phase) > 0:
                    response = messagebox.askyesno("選擇", "您要合併電流與相位的圖表嗎？")

                    if response:
                        # 建立分頁
                        wb = Workbook() # 刷新工作表
                        ws = wb.active
                        ws.title = "Current & Phase"
                        
                        # 合併電流與相位資料
                        all_data = []
                        all_data, data_filled = Merge_data(Current, Phase)
                        
                        # 儲存點資料
                        for row in all_data:
                            ws.append(row)

                        # 儲存掃頻最大點
                        if self.var_max.get():
                            # 掃頻最大點合併
                            if len(UD2_Current) > 0 and len(UD7_Current) > 0:
                                max_point = UD7_max_point + UD2_max_point[1:]
                                min_point = UD7_min_point + UD2_min_point[1:]
                            
                            elif len(UD2_Current) > 0 and len(UD7_Current) == 0:       
                                for i5 in range(len(UD2_Current[0]) - 1):
                                    UD2_max_point.insert(1, "") 
                                    UD2_min_point.insert(1, "")
                                
                                max_point = UD2_max_point
                                min_point = UD2_min_point
 
                            ws.append(max_point)
                            ws.append(min_point)
                        
                        # 儲存圖表
                        if len(all_data[0]) <= 256:
                            R = [ [2, len(Current[0]) + 1, len(Current[0]) - 1],
                                  [len(Current[0]) + 1, len(Current[0]) + len(Phase[0]), len(Phase[0]) - 1] ]
                            chart, adress, xMax, xMin = Drawing("Current & Phase", "", all_data, Lcolors, Ltypes, Lwidth, ws, R)
                            ws.add_chart(chart, adress)
                        
                        else:
                            messagebox.showwarning("圖表錯誤", "繪製圖表發生錯誤：\n\nExcel一張圖表最多只能容納 255 個資料數列！")
                            
                # 將資料存進Excel
                if self.var_c.get() or self.var_p.get():
                    try:
                        save_path = f'{self.latest_folder}/{d_type}_Output.xlsx'
                        wb.save(save_path)

                        # 資料空點圖表補線
                        if data_filled:
                            set_excel_chart_options(save_path)
                             
                        # 確認檔案是否存在的成功訊息
                        messagebox.showinfo("成功", f"成功合併檔案：{d_type}_Output.xlsx")
                        
                    except Exception as e:
                        print(f"⚠️檢查 1：讀確認{d_type}_Output.xlsx，檔案是否有開啟！")
                        messagebox.showerror("錯誤", f"儲存檔案時，發生錯誤：\n{str(e)}\n\n請關閉Excel，再試一次！")
                    
                else:
                    print("⚠️注意：未選擇任何合併資料！")
                    messagebox.showwarning("錯誤", "未選擇任何合併資料！")
  
            except Exception as e:
                print("錯誤原因：", e)
                print("⚠️檢查：讀確認選擇的檔案，為驅動器的掃頻資料！")
                messagebox.showerror("錯誤", "⚠️讀確認選擇的檔案，為驅動器的掃頻資料！")
                
        else:
            print("⚠️警告：檔案數量 ≠ 刀把數量 x 刀具數量")
            messagebox.showwarning("錯誤", "檔案數量不匹配，合併檔案失敗！！")

    # 離開
    def close_action(self):
        root.destroy()

# 運行主循環
if __name__ == "__main__":
    root = tk.Tk()
    app = UD_App(root)
    root.mainloop()